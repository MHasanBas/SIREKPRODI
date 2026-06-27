import argparse
import contextlib
import itertools
import io
import json
import math
import os
import pickle
import statistics
import time
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd

from clustering.kmeans_module import _run_kmeans_single


DEFAULT_ACTIVE_MODEL_FILE = "models/active_model.txt"
DEFAULT_OUTPUT_ROOT = "outputs/ga_tuning"
TUNING_PRESETS = {
    "quick": {
        "seeds": "42,43",
        "pop_sizes": "10,20",
        "generations": "10",
        "early_mutation_rates": "0.2,0.28",
        "mid_mutation_rates": "0.1,0.14",
        "late_mutation_rates": "0.01,0.05",
        "max_stagnants": "8",
    },
    "balanced": {
        "seeds": "42,43,44",
        "pop_sizes": "10,20",
        "generations": "10,25",
        "early_mutation_rates": "0.2,0.28",
        "mid_mutation_rates": "0.1,0.14",
        "late_mutation_rates": "0.01,0.05",
        "max_stagnants": "8",
    },
    "full": {
        "seeds": "42,43,44,45,46",
        "pop_sizes": "10,20,30",
        "generations": "10,25,50",
        "early_mutation_rates": "0.2,0.28,0.35",
        "mid_mutation_rates": "0.1,0.14,0.2",
        "late_mutation_rates": "0.01,0.05,0.1",
        "max_stagnants": "5,8,12",
    },
}


def parse_int_list(raw_value):
    return [int(part.strip()) for part in raw_value.split(",") if part.strip()]


def parse_float_list(raw_value):
    return [float(part.strip()) for part in raw_value.split(",") if part.strip()]


def load_dataframe(args):
    if args.data_pkl:
        data_path = args.data_pkl
    else:
        with open(args.active_model_file, "r", encoding="utf-8") as f:
            active_model = f.read().strip()
        if not active_model:
            raise ValueError(f"File model aktif kosong: {args.active_model_file}")
        data_path = os.path.join("models", active_model, "data_gabungan_clean.pkl")

    if not os.path.exists(data_path):
        raise FileNotFoundError(f"Data tidak ditemukan: {data_path}")

    with open(data_path, "rb") as f:
        payload = pickle.load(f)

    df = payload["data"] if isinstance(payload, dict) and "data" in payload else payload
    if not isinstance(df, pd.DataFrame):
        raise TypeError(f"Payload pada {data_path} bukan DataFrame yang valid.")

    return df.copy(), data_path


def make_output_dir(output_root):
    timestamp = datetime.now(ZoneInfo("Asia/Jakarta")).strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(output_root, f"run_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def safe_mean(values):
    valid = [v for v in values if v is not None and math.isfinite(v)]
    if not valid:
        return None
    return float(statistics.mean(valid))


def safe_std(values):
    valid = [v for v in values if v is not None and math.isfinite(v)]
    if len(valid) < 2:
        return 0.0 if valid else None
    return float(statistics.pstdev(valid))


def rounded(value, digits=4):
    if value is None:
        return None
    return round(value, digits)


def validate_rates(label, values):
    invalid = [value for value in values if value < 0 or value > 1]
    if invalid:
        raise ValueError(f"{label} harus berada pada rentang 0..1. Nilai tidak valid: {invalid}")


def apply_preset(args):
    if (
        args.mutation_rates
        and args.early_mutation_rates is None
        and args.mid_mutation_rates is None
        and args.late_mutation_rates is None
    ):
        # Classic GA tuning mode: one mutation_rate value is applied to all phases.
        args.early_mutation_rates = args.mutation_rates
        args.mid_mutation_rates = args.mutation_rates
        args.late_mutation_rates = args.mutation_rates
        args.mutation_mode = "single_rate_all_phases"
    else:
        args.mutation_mode = "phase_schedule"

    preset = TUNING_PRESETS[args.preset]
    for key, value in preset.items():
        if getattr(args, key) is None:
            setattr(args, key, value)

    if args.mutation_rates and args.mutation_mode != "single_rate_all_phases":
        args.late_mutation_rates = args.mutation_rates
    return args


def rank_summary(df_summary):
    if df_summary.empty:
        return df_summary

    ranked = df_summary.sort_values(
        by=[
            "mean_dbi_after_ga",
            "mean_silhouette",
            "mean_dbi_improvement_pct",
            "mean_ch_score",
            "mean_runtime_seconds",
        ],
        ascending=[True, False, False, False, True],
        na_position="last",
    ).reset_index(drop=True)
    ranked.insert(0, "rank", range(1, len(ranked) + 1))
    return ranked


def run_tuning(df, args, output_dir, progress_callback=None):
    seeds = parse_int_list(args.seeds)
    pop_sizes = parse_int_list(args.pop_sizes)
    generations_list = parse_int_list(args.generations)
    early_mutation_rates = parse_float_list(args.early_mutation_rates)
    mid_mutation_rates = parse_float_list(args.mid_mutation_rates)
    late_mutation_rates = parse_float_list(args.late_mutation_rates)
    max_stagnants = parse_int_list(args.max_stagnants)

    validate_rates("early_mutation_rates", early_mutation_rates)
    validate_rates("mid_mutation_rates", mid_mutation_rates)
    validate_rates("late_mutation_rates", late_mutation_rates)

    if args.mutation_mode == "single_rate_all_phases":
        mutation_schedules = [(rate, rate, rate) for rate in late_mutation_rates]
    else:
        mutation_schedules = list(itertools.product(
            early_mutation_rates,
            mid_mutation_rates,
            late_mutation_rates,
        ))

    combinations = [
        (pop_size, generations, early_rate, mid_rate, late_rate, max_stagnant)
        for pop_size, generations, (early_rate, mid_rate, late_rate), max_stagnant
        in itertools.product(pop_sizes, generations_list, mutation_schedules, max_stagnants)
    ]
    total_runs = len(combinations) * len(seeds)

    print(f"Dataset: {len(df)} baris")
    print(f"Mode mutation rate: {args.mutation_mode}")
    print(f"Kombinasi hyperparameter: {len(combinations)}")
    print(f"Jumlah seed per kombinasi: {len(seeds)}")
    print(f"Total eksekusi: {total_runs}")
    print(f"Output directory: {output_dir}")

    per_run_rows = []
    summary_rows = []
    completed_runs = 0

    for combo_index, (
        pop_size,
        generations,
        early_mutation_rate,
        mid_mutation_rate,
        late_mutation_rate,
        max_stagnant,
    ) in enumerate(combinations, start=1):
        combo_label = (
            f"pop={pop_size}, gen={generations}, "
            f"mut=({early_mutation_rate}/{mid_mutation_rate}/{late_mutation_rate}), "
            f"stagnant={max_stagnant}"
        )
        print(f"\n[{combo_index}/{len(combinations)}] Menjalankan {combo_label}")

        combo_results = []
        combo_runtimes = []

        for seed in seeds:
            start = time.perf_counter()
            run_kwargs = {
                "df": df,
                "n_clusters": args.n_clusters,
                "random_seed": seed,
                "ga_pop_size": pop_size,
                "ga_generations": generations,
                "ga_mutation_rate": late_mutation_rate,
                "ga_max_stagnant": max_stagnant,
                "ga_early_mutation_rate": early_mutation_rate,
                "ga_mid_mutation_rate": mid_mutation_rate,
                "ga_late_mutation_rate": late_mutation_rate,
            }
            if args.verbose:
                result = _run_kmeans_single(**run_kwargs)
            else:
                with contextlib.redirect_stdout(io.StringIO()):
                    result = _run_kmeans_single(**run_kwargs)
            runtime_seconds = time.perf_counter() - start
            combo_results.append(result)
            combo_runtimes.append(runtime_seconds)
            completed_runs += 1

            per_run_rows.append({
                "population_size": pop_size,
                "generations": generations,
                "early_mutation_rate": early_mutation_rate,
                "mid_mutation_rate": mid_mutation_rate,
                "late_mutation_rate": late_mutation_rate,
                "mutation_rate": late_mutation_rate,
                "max_stagnant": max_stagnant,
                "seed": seed,
                "runtime_seconds": round(runtime_seconds, 4),
                "dbi_before_ga": result.get("dbi_before_ga"),
                "dbi_after_ga": result.get("dbi_after_ga"),
                "dbi_improvement_pct": result.get("dbi_improvement_pct"),
                "silhouette_score": result.get("silhouette_score"),
                "ch_score": result.get("ch_score"),
                "total_inertia": result.get("total_inertia"),
                "final_k": result.get("final_k"),
                "n_clusters_used": result.get("n_clusters_used"),
            })
            print(
                f"  Seed {seed} selesai "
                f"({completed_runs}/{total_runs}) | "
                f"DBI={result.get('dbi_after_ga')} | "
                f"Silhouette={result.get('silhouette_score')} | "
                f"{runtime_seconds:.2f}s"
            )
            if progress_callback:
                progress_callback(
                    completed_runs,
                    total_runs,
                    {
                        "combo_index": combo_index,
                        "combo_total": len(combinations),
                        "seed": seed,
                        "dbi_after_ga": result.get("dbi_after_ga"),
                        "silhouette_score": result.get("silhouette_score"),
                        "runtime_seconds": runtime_seconds,
                        "label": combo_label,
                    },
                )

        summary_rows.append({
            "population_size": pop_size,
            "generations": generations,
            "early_mutation_rate": early_mutation_rate,
            "mid_mutation_rate": mid_mutation_rate,
            "late_mutation_rate": late_mutation_rate,
            "mutation_rate": late_mutation_rate,
            "max_stagnant": max_stagnant,
            "n_runs": len(combo_results),
            "mean_runtime_seconds": rounded(safe_mean(combo_runtimes), 4),
            "std_runtime_seconds": rounded(safe_std(combo_runtimes), 4),
            "mean_dbi_before_ga": rounded(safe_mean([r.get("dbi_before_ga") for r in combo_results]), 4),
            "std_dbi_before_ga": rounded(safe_std([r.get("dbi_before_ga") for r in combo_results]), 4),
            "mean_dbi_after_ga": rounded(safe_mean([r.get("dbi_after_ga") for r in combo_results]), 4),
            "std_dbi_after_ga": rounded(safe_std([r.get("dbi_after_ga") for r in combo_results]), 4),
            "mean_dbi_improvement_pct": rounded(safe_mean([r.get("dbi_improvement_pct") for r in combo_results]), 4),
            "std_dbi_improvement_pct": rounded(safe_std([r.get("dbi_improvement_pct") for r in combo_results]), 4),
            "mean_silhouette": rounded(safe_mean([r.get("silhouette_score") for r in combo_results]), 4),
            "std_silhouette": rounded(safe_std([r.get("silhouette_score") for r in combo_results]), 4),
            "mean_ch_score": rounded(safe_mean([r.get("ch_score") for r in combo_results]), 4),
            "std_ch_score": rounded(safe_std([r.get("ch_score") for r in combo_results]), 4),
            "best_single_run_dbi": min(
                [r.get("dbi_after_ga") for r in combo_results if r.get("dbi_after_ga") is not None],
                default=None,
            ),
            "best_single_run_silhouette": max(
                [r.get("silhouette_score") for r in combo_results if r.get("silhouette_score") is not None],
                default=None,
            ),
        })

    return pd.DataFrame(per_run_rows), rank_summary(pd.DataFrame(summary_rows))


def build_manifest(args, data_path, output_dir, n_data_rows, per_run_df, summary_df):
    best_config = summary_df.iloc[0].to_dict() if not summary_df.empty else None
    manifest = {
        "created_at": datetime.now(ZoneInfo("Asia/Jakarta")).isoformat(timespec="seconds"),
        "data_path": data_path,
        "n_data_rows": int(n_data_rows),
        "n_executions": int(len(per_run_df.index)) if not per_run_df.empty else 0,
        "n_clusters": args.n_clusters,
        "seeds": parse_int_list(args.seeds),
        "pop_sizes": parse_int_list(args.pop_sizes),
        "generations": parse_int_list(args.generations),
        "early_mutation_rates": parse_float_list(args.early_mutation_rates),
        "mid_mutation_rates": parse_float_list(args.mid_mutation_rates),
        "late_mutation_rates": parse_float_list(args.late_mutation_rates),
        "mutation_mode": args.mutation_mode,
        "max_stagnants": parse_int_list(args.max_stagnants),
        "output_dir": output_dir,
        "notes": [
            "Pipeline tuning memakai _run_kmeans_single dari clustering.kmeans_module.",
            "Adaptive mutation schedule dituning eksplisit per fase: early, mid, dan late mutation rate.",
            "Baseline KMeans di pipeline memakai random_state tetap 42, bukan seed tuning per run.",
        ],
        "best_config": best_config,
        "files": {
            "per_run_csv": os.path.join(output_dir, "ga_tuning_per_run.csv"),
            "summary_csv": os.path.join(output_dir, "ga_tuning_summary.csv"),
            "summary_xlsx": os.path.join(output_dir, "ga_tuning_results.xlsx"),
            "manifest_json": os.path.join(output_dir, "ga_tuning_manifest.json"),
        },
    }
    return manifest


def export_thesis_excel(summary_df, per_run_df, output_path, args, data_path, n_data_rows):
    """
    Ekspor hasil tuning ke Excel multi-sheet yang komprehensif dan siap pakai untuk skripsi.
    Mencakup: ranking, top configs, sensitivitas parameter, per-run detail, pivot impact,
    dan sheet ringkasan metodologi.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError:
        print("  Peringatan: openpyxl tidak tersedia, skip thesis Excel export.")
        return

    # ── Warna Palet ────────────────────────────────────────────────────────────
    CLR_HEADER_DARK   = "1A3A5C"   # Biru gelap — header utama
    CLR_HEADER_MED    = "2E6DA4"   # Biru medium — sub-header
    CLR_HEADER_LIGHT  = "D6E4F0"   # Biru muda — header biasa
    CLR_GOLD_TOP      = "FFF2CC"   # Kuning emas — rank 1–5
    CLR_GREEN_GOOD    = "E2EFDA"   # Hijau — rank 6–20
    CLR_WHITE_EVEN    = "FFFFFF"
    CLR_WHITE_ODD     = "F7F9FC"
    CLR_ACCENT_BLUE   = "BDD7EE"
    CLR_RED_POOR      = "FCE4EC"   # Merah muda — hasil buruk
    CLR_BORDER        = "B0BEC5"
    CLR_BEST_DBI      = "00C853"   # Hijau terang — nilai DBI terbaik

    def _fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _font(bold=False, size=11, color="000000", italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

    def _border_thin():
        s = Side(border_style="thin", color=CLR_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _apply_header_row(ws, row_num, labels, col_widths=None, bg=CLR_HEADER_DARK, fg="FFFFFF", size=11):
        for col_idx, label in enumerate(labels, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=label)
            cell.fill = _fill(bg)
            cell.font = _font(bold=True, color=fg, size=size)
            cell.alignment = _center()
            cell.border = _border_thin()
            if col_widths and col_idx <= len(col_widths):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    def _style_data_row(ws, row_num, n_cols, fill_color, bold=False):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = _fill(fill_color)
            cell.font = _font(bold=bold)
            cell.alignment = _left()
            cell.border = _border_thin()

    def _fmt_pct(v):
        return f"{v:+.2f}%" if v is not None else "-"

    def _fmt_num(v, decimals=4):
        if v is None:
            return "-"
        return round(float(v), decimals)

    # Pastikan summary_df ada rank
    if "rank" not in summary_df.columns:
        summary_df = summary_df.copy()
        summary_df.insert(0, "rank", range(1, len(summary_df) + 1))

    best_dbi = summary_df["mean_dbi_after_ga"].min() if not summary_df.empty else None

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1: 🏆 Ranking Lengkap
        # ══════════════════════════════════════════════════════════════════════
        headers_rank = [
            "Rank", "Pop Size", "Generasi", "Early Mut Rate",
            "Mid Mut Rate", "Late Mut Rate", "Max Stagnant",
            "Avg DBI Sebelum GA", "Avg DBI Setelah GA ↓", "Std DBI",
            "Avg Silhouette ↑", "Std Silhouette",
            "Avg CH Score ↑", "Avg Impr DBI (%)",
            "Best Single DBI", "Avg Runtime (s)", "Ket."
        ]
        col_widths_rank = [7, 10, 10, 15, 14, 14, 13, 18, 18, 10, 17, 14, 14, 17, 15, 15, 30]

        sheet1_rows = []
        for _, r in summary_df.iterrows():
            dbi_ga = r.get("mean_dbi_after_ga")
            impr = r.get("mean_dbi_improvement_pct")
            is_best = (dbi_ga is not None and best_dbi is not None and abs(float(dbi_ga) - float(best_dbi)) < 1e-6)
            ket = "✅ TERBAIK" if is_best else ("⭐ Top 5" if int(r.get("rank", 99)) <= 5 else "")
            sheet1_rows.append({
                "Rank": int(r.get("rank", 0)),
                "Pop Size": int(r.get("population_size", 0)),
                "Generasi": int(r.get("generations", 0)),
                "Early Mut Rate": _fmt_num(r.get("early_mutation_rate"), 2),
                "Mid Mut Rate": _fmt_num(r.get("mid_mutation_rate"), 2),
                "Late Mut Rate": _fmt_num(r.get("late_mutation_rate"), 2),
                "Max Stagnant": int(r.get("max_stagnant", 0)),
                "Avg DBI Sebelum GA": _fmt_num(r.get("mean_dbi_before_ga")),
                "Avg DBI Setelah GA ↓": _fmt_num(dbi_ga),
                "Std DBI": _fmt_num(r.get("std_dbi_after_ga")),
                "Avg Silhouette ↑": _fmt_num(r.get("mean_silhouette")),
                "Std Silhouette": _fmt_num(r.get("std_silhouette")),
                "Avg CH Score ↑": _fmt_num(r.get("mean_ch_score"), 2),
                "Avg Impr DBI (%)": _fmt_num(impr, 2),
                "Best Single DBI": _fmt_num(r.get("best_single_run_dbi")),
                "Avg Runtime (s)": _fmt_num(r.get("mean_runtime_seconds"), 2),
                "Ket.": ket,
            })

        df_sheet1 = pd.DataFrame(sheet1_rows)
        df_sheet1.to_excel(writer, sheet_name="🏆 Ranking Lengkap", index=False)
        ws1 = writer.sheets["🏆 Ranking Lengkap"]

        # Style header
        _apply_header_row(ws1, 1, headers_rank, col_widths_rank, bg=CLR_HEADER_DARK)
        ws1.row_dimensions[1].height = 36

        # Style baris data
        for row_idx, (_, r) in enumerate(df_sheet1.iterrows(), start=2):
            rank_val = r.get("Rank", 99)
            if rank_val == 1:
                row_color = CLR_GOLD_TOP
                bold = True
            elif rank_val <= 5:
                row_color = "FFF9C4"
                bold = False
            elif rank_val <= 20:
                row_color = CLR_GREEN_GOOD
                bold = False
            elif row_idx % 2 == 0:
                row_color = CLR_WHITE_EVEN
                bold = False
            else:
                row_color = CLR_WHITE_ODD
                bold = False
            _style_data_row(ws1, row_idx, len(headers_rank), row_color, bold=bold)
            ws1.row_dimensions[row_idx].height = 18

        # Freeze panes
        ws1.freeze_panes = "A2"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2: 🥇 Top 20 Konfigurasi Terbaik
        # ══════════════════════════════════════════════════════════════════════
        top20 = summary_df.head(20).copy()
        top20_rows = []
        for _, r in top20.iterrows():
            dbi_ga = r.get("mean_dbi_after_ga")
            dbi_before = r.get("mean_dbi_before_ga")
            impr = r.get("mean_dbi_improvement_pct")
            top20_rows.append({
                "Rank": int(r.get("rank", 0)),
                "Konfigurasi": (
                    f"Pop={int(r.get('population_size',0))}, "
                    f"Gen={int(r.get('generations',0))}, "
                    f"EMut={r.get('early_mutation_rate','-')}, "
                    f"MMut={r.get('mid_mutation_rate','-')}, "
                    f"LMut={r.get('late_mutation_rate','-')}, "
                    f"Stag={int(r.get('max_stagnant',0))}"
                ),
                "Pop Size": int(r.get("population_size", 0)),
                "Generasi": int(r.get("generations", 0)),
                "Early Mut": _fmt_num(r.get("early_mutation_rate"), 2),
                "Mid Mut": _fmt_num(r.get("mid_mutation_rate"), 2),
                "Late Mut": _fmt_num(r.get("late_mutation_rate"), 2),
                "Max Stagnant": int(r.get("max_stagnant", 0)),
                "DBI Baseline": _fmt_num(dbi_before),
                "DBI Setelah GA": _fmt_num(dbi_ga),
                "Std DBI": _fmt_num(r.get("std_dbi_after_ga")),
                "Peningkatan DBI": _fmt_num(impr, 2),
                "Silhouette": _fmt_num(r.get("mean_silhouette")),
                "CH Score": _fmt_num(r.get("mean_ch_score"), 2),
                "Best Run DBI": _fmt_num(r.get("best_single_run_dbi")),
                "Runtime (s)": _fmt_num(r.get("mean_runtime_seconds"), 1),
                "Interpretasi DBI": (
                    "✅ Sangat Baik (< 0.50)" if dbi_ga is not None and float(dbi_ga) < 0.50 else
                    "🟡 Baik (0.50–0.55)" if dbi_ga is not None and float(dbi_ga) < 0.55 else
                    "🟠 Cukup (0.55–0.60)" if dbi_ga is not None and float(dbi_ga) < 0.60 else
                    "🔴 Perlu Perbaikan"
                ),
            })

        df_top20 = pd.DataFrame(top20_rows)
        df_top20.to_excel(writer, sheet_name="🥇 Top 20 Konfigurasi", index=False)
        ws2 = writer.sheets["🥇 Top 20 Konfigurasi"]

        h2 = list(df_top20.columns)
        cw2 = [7, 60, 10, 10, 10, 10, 10, 13, 14, 14, 10, 15, 13, 12, 14, 12, 24]
        _apply_header_row(ws2, 1, h2, cw2, bg=CLR_HEADER_MED)
        ws2.row_dimensions[1].height = 36

        for row_idx, (_, r) in enumerate(df_top20.iterrows(), start=2):
            rank_val = r.get("Rank", 99)
            if rank_val == 1:
                row_color = CLR_GOLD_TOP
                bold = True
            elif rank_val <= 3:
                row_color = "FFF9C4"
                bold = False
            elif rank_val <= 10:
                row_color = CLR_GREEN_GOOD
                bold = False
            else:
                row_color = CLR_WHITE_ODD if row_idx % 2 else CLR_WHITE_EVEN
                bold = False
            _style_data_row(ws2, row_idx, len(h2), row_color, bold=bold)
            ws2.row_dimensions[row_idx].height = 20

        ws2.freeze_panes = "A2"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3: 📊 Sensitivitas Parameter
        # ══════════════════════════════════════════════════════════════════════
        param_cols = {
            "Population Size": "population_size",
            "Generasi": "generations",
            "Early Mutation Rate": "early_mutation_rate",
            "Mid Mutation Rate": "mid_mutation_rate",
            "Late Mutation Rate": "late_mutation_rate",
            "Max Stagnant": "max_stagnant",
        }

        ws3 = writer.book.create_sheet("📊 Sensitivitas Parameter")
        writer.sheets["📊 Sensitivitas Parameter"] = ws3

        current_row = 1

        # Judul
        ws3.merge_cells(f"A{current_row}:H{current_row}")
        title_cell = ws3.cell(row=current_row, column=1,
                              value="ANALISIS SENSITIVITAS HYPERPARAMETER TERHADAP DBI")
        title_cell.fill = _fill(CLR_HEADER_DARK)
        title_cell.font = _font(bold=True, color="FFFFFF", size=13)
        title_cell.alignment = _center()
        ws3.row_dimensions[current_row].height = 30
        current_row += 1

        ws3.merge_cells(f"A{current_row}:H{current_row}")
        sub_cell = ws3.cell(row=current_row, column=1,
                            value="Rata-rata DBI per nilai parameter — nilai DBI lebih kecil = clustering lebih baik")
        sub_cell.fill = _fill(CLR_HEADER_LIGHT)
        sub_cell.font = _font(italic=True, color="1A3A5C", size=10)
        sub_cell.alignment = _center()
        current_row += 2

        for param_label, col_name in param_cols.items():
            if col_name not in summary_df.columns:
                continue

            grp = summary_df.groupby(col_name).agg(
                jumlah_konfigurasi=("mean_dbi_after_ga", "count"),
                avg_dbi=("mean_dbi_after_ga", "mean"),
                std_dbi=("mean_dbi_after_ga", "std"),
                min_dbi=("mean_dbi_after_ga", "min"),
                avg_silhouette=("mean_silhouette", "mean"),
                avg_impr_pct=("mean_dbi_improvement_pct", "mean"),
            ).reset_index()

            grp_best_idx = grp["avg_dbi"].idxmin() if not grp.empty else None

            # Sub-header parameter
            ws3.merge_cells(f"A{current_row}:H{current_row}")
            ph = ws3.cell(row=current_row, column=1, value=f"▶  Parameter: {param_label}")
            ph.fill = _fill(CLR_HEADER_MED)
            ph.font = _font(bold=True, color="FFFFFF", size=11)
            ph.alignment = _left()
            ws3.row_dimensions[current_row].height = 22
            current_row += 1

            # Header kolom
            h3_labels = [
                param_label, "Jml Konfigurasi", "Avg DBI Setelah GA ↓",
                "Std DBI", "Min DBI (Best)", "Avg Silhouette ↑",
                "Avg Impr DBI (%)", "Status"
            ]
            _apply_header_row(ws3, current_row, h3_labels,
                              [18, 17, 22, 12, 16, 18, 18, 20],
                              bg=CLR_HEADER_LIGHT, fg=CLR_HEADER_DARK, size=10)
            ws3.row_dimensions[current_row].height = 20
            current_row += 1

            for grp_idx, row in grp.iterrows():
                is_best_param = (grp_best_idx is not None and grp_idx == grp_best_idx)
                row_color = CLR_GOLD_TOP if is_best_param else (CLR_WHITE_ODD if (current_row % 2) else CLR_WHITE_EVEN)
                avg_dbi_val = row.get("avg_dbi")
                cells_data = [
                    row[col_name],
                    int(row.get("jumlah_konfigurasi", 0)),
                    _fmt_num(avg_dbi_val),
                    _fmt_num(row.get("std_dbi")),
                    _fmt_num(row.get("min_dbi")),
                    _fmt_num(row.get("avg_silhouette")),
                    _fmt_num(row.get("avg_impr_pct"), 2),
                    "✅ TERBAIK" if is_best_param else "",
                ]
                for col_idx, val in enumerate(cells_data, start=1):
                    cell = ws3.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = _fill(row_color)
                    cell.font = _font(bold=is_best_param)
                    cell.alignment = _center()
                    cell.border = _border_thin()
                ws3.row_dimensions[current_row].height = 16
                current_row += 1

            current_row += 1  # Spasi antar parameter

        # Set lebar kolom ws3
        for col_letter, width in zip(["A","B","C","D","E","F","G","H"],
                                     [18, 17, 22, 12, 16, 18, 18, 20]):
            ws3.column_dimensions[col_letter].width = width

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4: 📋 Per-Run Detail
        # ══════════════════════════════════════════════════════════════════════
        df_per_run_sorted = per_run_df.sort_values("dbi_after_ga", ascending=True).reset_index(drop=True)
        df_per_run_sorted.insert(0, "no", range(1, len(df_per_run_sorted) + 1))

        friendly_cols = {
            "no": "No.",
            "population_size": "Pop Size",
            "generations": "Generasi",
            "early_mutation_rate": "Early Mut",
            "mid_mutation_rate": "Mid Mut",
            "late_mutation_rate": "Late Mut",
            "max_stagnant": "Max Stagnant",
            "seed": "Seed",
            "runtime_seconds": "Runtime (s)",
            "dbi_before_ga": "DBI Baseline",
            "dbi_after_ga": "DBI Setelah GA ↓",
            "dbi_improvement_pct": "Impr DBI (%)",
            "silhouette_score": "Silhouette ↑",
            "ch_score": "CH Score ↑",
            "total_inertia": "Total Inertia",
            "final_k": "K Cluster",
        }
        cols_to_export = [c for c in friendly_cols if c in df_per_run_sorted.columns]
        df_per_export = df_per_run_sorted[cols_to_export].rename(columns=friendly_cols)
        df_per_export.to_excel(writer, sheet_name="📋 Per-Run Detail", index=False)
        ws4 = writer.sheets["📋 Per-Run Detail"]

        per_headers = list(df_per_export.columns)
        per_widths   = [7, 10, 10, 10, 10, 10, 13, 8, 12, 15, 17, 13, 14, 12, 14, 10]
        _apply_header_row(ws4, 1, per_headers, per_widths, bg=CLR_HEADER_DARK)
        ws4.row_dimensions[1].height = 30

        best_per_dbi = df_per_export["DBI Setelah GA ↓"].min() if "DBI Setelah GA ↓" in df_per_export.columns else None
        for row_idx in range(2, len(df_per_export) + 2):
            cell_dbi = ws4.cell(row=row_idx, column=per_headers.index("DBI Setelah GA ↓") + 1) if "DBI Setelah GA ↓" in per_headers else None
            dbi_val = cell_dbi.value if cell_dbi else None
            is_best_run = (dbi_val is not None and best_per_dbi is not None
                           and abs(float(dbi_val) - float(best_per_dbi)) < 1e-6)
            row_color = CLR_GOLD_TOP if is_best_run else (CLR_WHITE_ODD if row_idx % 2 else CLR_WHITE_EVEN)
            _style_data_row(ws4, row_idx, len(per_headers), row_color, bold=is_best_run)
            ws4.row_dimensions[row_idx].height = 15

        ws4.freeze_panes = "A2"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5: 📈 Pivot: Impact Parameter vs DBI
        # ══════════════════════════════════════════════════════════════════════
        ws5 = writer.book.create_sheet("📈 Impact Parameter vs DBI")
        writer.sheets["📈 Impact Parameter vs DBI"] = ws5

        ws5.merge_cells("A1:F1")
        t5 = ws5.cell(row=1, column=1, value="DAMPAK SETIAP PARAMETER TERHADAP RATA-RATA DBI SETELAH GA")
        t5.fill = _fill(CLR_HEADER_DARK)
        t5.font = _font(bold=True, color="FFFFFF", size=13)
        t5.alignment = _center()
        ws5.row_dimensions[1].height = 28

        ws5.merge_cells("A2:F2")
        s5 = ws5.cell(row=2, column=1,
                      value="Tabel ini berguna untuk menentukan parameter mana yang paling berpengaruh terhadap kualitas clustering.")
        s5.fill = _fill(CLR_HEADER_LIGHT)
        s5.font = _font(italic=True, size=10)
        s5.alignment = _center()

        cr5 = 4
        pivot_data = []
        for param_label, col_name in param_cols.items():
            if col_name not in summary_df.columns:
                continue
            for val, grp in summary_df.groupby(col_name):
                avg_d = grp["mean_dbi_after_ga"].mean()
                min_d = grp["mean_dbi_after_ga"].min()
                count = len(grp)
                pivot_data.append({
                    "Parameter": param_label,
                    "Nilai Parameter": val,
                    "Jumlah Kombinasi": count,
                    "Avg DBI Setelah GA": round(avg_d, 4) if not pd.isna(avg_d) else None,
                    "Min DBI (Best Run)": round(min_d, 4) if not pd.isna(min_d) else None,
                    "Selisih vs Global Min": None,
                })

        global_min_dbi = summary_df["mean_dbi_after_ga"].min() if not summary_df.empty else None
        for row in pivot_data:
            if row["Avg DBI Setelah GA"] is not None and global_min_dbi is not None:
                row["Selisih vs Global Min"] = round(float(row["Avg DBI Setelah GA"]) - float(global_min_dbi), 4)

        df_pivot = pd.DataFrame(pivot_data)
        h5 = list(df_pivot.columns)
        cw5 = [22, 18, 18, 22, 18, 22]
        _apply_header_row(ws5, cr5, h5, cw5, bg=CLR_HEADER_MED)
        ws5.row_dimensions[cr5].height = 24
        cr5 += 1

        prev_param = None
        for _, pr in df_pivot.iterrows():
            is_new_param = pr["Parameter"] != prev_param
            if is_new_param:
                prev_param = pr["Parameter"]
                row_color = CLR_ACCENT_BLUE
            else:
                row_color = CLR_WHITE_ODD if cr5 % 2 else CLR_WHITE_EVEN

            selisih = pr.get("Selisih vs Global Min")
            is_global_best = selisih is not None and abs(float(selisih)) < 1e-6

            for col_idx, key in enumerate(h5, start=1):
                cell = ws5.cell(row=cr5, column=col_idx, value=pr[key])
                cell.fill = _fill(CLR_GOLD_TOP if is_global_best else row_color)
                cell.font = _font(bold=is_global_best)
                cell.alignment = _center()
                cell.border = _border_thin()
            ws5.row_dimensions[cr5].height = 16
            cr5 += 1

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 6: 📝 Ringkasan Metodologi (untuk Skripsi)
        # ══════════════════════════════════════════════════════════════════════
        ws6 = writer.book.create_sheet("📝 Ringkasan Metodologi")
        writer.sheets["📝 Ringkasan Metodologi"] = ws6

        ws6.column_dimensions["A"].width = 35
        ws6.column_dimensions["B"].width = 55
        ws6.column_dimensions["C"].width = 35

        def _meta_row(ws, row, label, value, note="", label_bg=CLR_HEADER_LIGHT):
            c1 = ws.cell(row=row, column=1, value=label)
            c1.fill = _fill(label_bg)
            c1.font = _font(bold=True, size=10, color=CLR_HEADER_DARK)
            c1.alignment = _left()
            c1.border = _border_thin()
            c2 = ws.cell(row=row, column=2, value=value)
            c2.fill = _fill(CLR_WHITE_EVEN)
            c2.font = _font(size=10)
            c2.alignment = _left()
            c2.border = _border_thin()
            c3 = ws.cell(row=row, column=3, value=note)
            c3.fill = _fill("FFFDE7")
            c3.font = _font(italic=True, size=9, color="5D4037")
            c3.alignment = _left()
            c3.border = _border_thin()
            ws.row_dimensions[row].height = 20

        ws6.merge_cells("A1:C1")
        t6 = ws6.cell(row=1, column=1, value="RINGKASAN EKSPERIMEN HYPERPARAMETER TUNING GA+KMEANS")
        t6.fill = _fill(CLR_HEADER_DARK)
        t6.font = _font(bold=True, color="FFFFFF", size=14)
        t6.alignment = _center()
        ws6.row_dimensions[1].height = 32

        ws6.merge_cells("A2:C2")
        ws6.cell(row=2, column=1, value="Referensi: Kolom A = Label, Kolom B = Nilai, Kolom C = Catatan/Interpretasi")
        ws6.cell(row=2, column=1).fill = _fill(CLR_HEADER_LIGHT)
        ws6.cell(row=2, column=1).font = _font(italic=True, size=9)
        ws6.cell(row=2, column=1).alignment = _center()

        # Block: Info Dataset
        ws6.merge_cells("A4:C4")
        h_ds = ws6.cell(row=4, column=1, value="A. INFORMASI DATASET")
        h_ds.fill = _fill(CLR_HEADER_MED)
        h_ds.font = _font(bold=True, color="FFFFFF", size=11)
        ws6.row_dimensions[4].height = 22

        _meta_row(ws6, 5, "Jumlah Data (baris)", n_data_rows, "Total mahasiswa dalam dataset")
        _meta_row(ws6, 6, "Sumber Data", data_path, "File pickle model aktif")
        _meta_row(ws6, 7, "Jumlah Kluster (K)", getattr(args, "n_clusters", 3), "K yang digunakan dalam eksperimen")

        # Block: Konfigurasi Grid
        ws6.merge_cells("A9:C9")
        h_gs = ws6.cell(row=9, column=1, value="B. RUANG PENCARIAN HYPERPARAMETER (GRID SEARCH)")
        h_gs.fill = _fill(CLR_HEADER_MED)
        h_gs.font = _font(bold=True, color="FFFFFF", size=11)
        ws6.row_dimensions[9].height = 22

        seeds_val = getattr(args, "seeds", "-")
        pop_val = getattr(args, "pop_sizes", "-")
        gen_val = getattr(args, "generations", "-")
        emut_val = getattr(args, "early_mutation_rates", "-")
        mmut_val = getattr(args, "mid_mutation_rates", "-")
        lmut_val = getattr(args, "late_mutation_rates", "-")
        stag_val = getattr(args, "max_stagnants", "-")
        n_total = len(per_run_df) if per_run_df is not None else 0
        n_combos = len(summary_df) if not summary_df.empty else 0

        _meta_row(ws6, 10, "Seeds (Random State)", seeds_val, "Untuk reproducibility; setiap kombinasi diuji per seed")
        _meta_row(ws6, 11, "Population Size", pop_val, "Jumlah individu dalam populasi GA per generasi")
        _meta_row(ws6, 12, "Generations", gen_val, "Jumlah iterasi evolusi GA")
        _meta_row(ws6, 13, "Early Mutation Rate", emut_val, "Fase awal: eksplorasi global (nilai tinggi = lebih eksploratif)")
        _meta_row(ws6, 14, "Mid Mutation Rate", mmut_val, "Fase tengah: transisi eksplorasi→eksploitasi")
        _meta_row(ws6, 15, "Late Mutation Rate", lmut_val, "Fase akhir: fine-tuning (nilai rendah = konvergensi stabil)")
        _meta_row(ws6, 16, "Max Stagnant", stag_val, "Batas generasi tanpa peningkatan sebelum early-stop")
        _meta_row(ws6, 17, "Total Kombinasi Grid", n_combos, "Jumlah konfigurasi unik yang diuji")
        _meta_row(ws6, 18, "Total Run Eksekusi", n_total, f"= {n_combos} kombinasi × {len(seeds_val.split(',')) if isinstance(seeds_val, str) else '?'} seeds")

        # Block: Hasil Terbaik
        ws6.merge_cells("A20:C20")
        h_res = ws6.cell(row=20, column=1, value="C. HASIL EKSPERIMEN & KONFIGURASI TERBAIK")
        h_res.fill = _fill(CLR_HEADER_MED)
        h_res.font = _font(bold=True, color="FFFFFF", size=11)
        ws6.row_dimensions[20].height = 22

        if not summary_df.empty:
            best_row = summary_df.iloc[0]
            _meta_row(ws6, 21, "Rank 1 — Population Size", int(best_row.get("population_size", "-")), "Jumlah individu populasi GA terbaik")
            _meta_row(ws6, 22, "Rank 1 — Generations", int(best_row.get("generations", "-")), "Jumlah generasi terbaik")
            _meta_row(ws6, 23, "Rank 1 — Early Mutation Rate", best_row.get("early_mutation_rate", "-"), "")
            _meta_row(ws6, 24, "Rank 1 — Mid Mutation Rate", best_row.get("mid_mutation_rate", "-"), "")
            _meta_row(ws6, 25, "Rank 1 — Late Mutation Rate", best_row.get("late_mutation_rate", "-"), "")
            _meta_row(ws6, 26, "Rank 1 — Max Stagnant", int(best_row.get("max_stagnant", "-")), "")
            _meta_row(ws6, 27, "DBI Baseline (Sebelum GA)", _fmt_num(best_row.get("mean_dbi_before_ga")),
                      "DBI KMeans standar tanpa optimasi GA")
            _meta_row(ws6, 28, "DBI Terbaik (Setelah GA)", _fmt_num(best_row.get("mean_dbi_after_ga")),
                      "DBI hasil optimasi GA — nilai lebih kecil lebih baik")
            _meta_row(ws6, 29, "Peningkatan DBI (%)",
                      f"{_fmt_num(best_row.get('mean_dbi_improvement_pct'), 2)}%",
                      "Persentase perbaikan DBI oleh GA vs KMeans standar")
            _meta_row(ws6, 30, "Silhouette Score Terbaik",
                      _fmt_num(best_row.get("mean_silhouette")),
                      "Mendekati 1 = cluster sangat terpisah & kohesif")
            _meta_row(ws6, 31, "Calinski-Harabasz Score",
                      _fmt_num(best_row.get("mean_ch_score"), 2),
                      "Nilai lebih besar = definisi cluster lebih tajam")
        else:
            ws6.cell(row=21, column=1, value="(Data tuning kosong)")

        # Block: Panduan Interpretasi
        ws6.merge_cells("A33:C33")
        h_int = ws6.cell(row=33, column=1, value="D. PANDUAN INTERPRETASI METRIK (UNTUK SKRIPSI)")
        h_int.fill = _fill(CLR_HEADER_MED)
        h_int.font = _font(bold=True, color="FFFFFF", size=11)
        ws6.row_dimensions[33].height = 22

        guides = [
            ("Davies-Bouldin Index (DBI)", "Mendekati 0 = sangat baik. Mengukur rata-rata kemiripan antar cluster. Semakin kecil = cluster semakin terpisah dan padat.", "Target: < 0.55 untuk dataset ini"),
            ("Silhouette Score", "Rentang −1 s/d 1. Mendekati 1 = cluster sangat baik; mendekati 0 = tumpang tindih; negatif = salah cluster.", "Target: > 0.50"),
            ("Calinski-Harabasz (CH) Score", "Tidak ada batas atas. Nilai lebih besar berarti cluster lebih terpisah dan kompak secara bersamaan.", "Bandingkan relatif antar eksperimen"),
            ("Early Mutation Rate", "Mengontrol eksplorasi di generasi awal GA. Nilai tinggi (0.28–0.35) mendorong diversifikasi populasi.", "Terlalu tinggi = chaos"),
            ("Mid Mutation Rate", "Transisi antara eksplorasi dan eksploitasi. Nilai menengah (0.10–0.14) adalah standar.", ""),
            ("Late Mutation Rate", "Fine-tuning di akhir evolusi. Nilai rendah (0.01–0.05) memastikan konvergensi stabil.", "Terlalu tinggi = divergen"),
            ("Max Stagnant", "Batas generasi tanpa peningkatan. Nilai besar (12) = lebih sabar; nilai kecil (5) = early stop agresif.", ""),
            ("Population Size", "Semakin besar = eksplorasi lebih luas tapi lebih lambat. Optimal: 20–40 untuk dataset ini.", ""),
            ("Generations", "Lebih banyak generasi = GA punya lebih banyak iterasi untuk konvergen. Diminishing returns setelah 50.", ""),
        ]
        for g_idx, (term, meaning, tip) in enumerate(guides, start=34):
            _meta_row(ws6, g_idx, term, meaning, tip)

        print(f"  ✅ Excel thesis-grade disimpan: {output_path}")


def main():

    parser = argparse.ArgumentParser(
        description="Grid search hyperparameter Genetic Algorithm untuk pipeline KMeans + GA."
    )
    parser.add_argument("--data-pkl", help="Path ke data_gabungan_clean.pkl. Default: model aktif.")
    parser.add_argument("--active-model-file", default=DEFAULT_ACTIVE_MODEL_FILE)
    parser.add_argument("--n-clusters", type=int, default=3)
    parser.add_argument("--preset", choices=sorted(TUNING_PRESETS), default="balanced")
    parser.add_argument("--seeds")
    parser.add_argument("--pop-sizes")
    parser.add_argument("--generations")
    parser.add_argument("--early-mutation-rates")
    parser.add_argument("--mid-mutation-rates")
    parser.add_argument("--late-mutation-rates")
    parser.add_argument("--mutation-rates", help="Deprecated: alias untuk --late-mutation-rates.")
    parser.add_argument("--max-stagnants")
    parser.add_argument("--output-root", default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--verbose", action="store_true", help="Tampilkan log detail setiap prodi selama tuning.")
    args = apply_preset(parser.parse_args())

    df, data_path = load_dataframe(args)
    output_dir = make_output_dir(args.output_root)

    per_run_df, summary_df = run_tuning(df, args, output_dir)

    per_run_csv = os.path.join(output_dir, "ga_tuning_per_run.csv")
    summary_csv = os.path.join(output_dir, "ga_tuning_summary.csv")
    summary_xlsx = os.path.join(output_dir, "ga_tuning_results.xlsx")
    manifest_json = os.path.join(output_dir, "ga_tuning_manifest.json")

    per_run_df.to_csv(per_run_csv, index=False)
    summary_df.to_csv(summary_csv, index=False)

    print("\nMembuat Excel thesis-grade (multi-sheet, color-coded)...")
    export_thesis_excel(summary_df, per_run_df, summary_xlsx, args, data_path, len(df))

    manifest = build_manifest(args, data_path, output_dir, len(df), per_run_df, summary_df)
    with open(manifest_json, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    print("\nTuning selesai.")
    print(f"Per-run CSV : {per_run_csv}")
    print(f"Summary CSV : {summary_csv}")
    print(f"Excel       : {summary_xlsx}")
    print(f"Manifest    : {manifest_json}")

    if not summary_df.empty:
        best = summary_df.iloc[0]
        print("\nKonfigurasi terbaik berdasarkan ranking summary:")
        print(
            f"rank={int(best['rank'])}, pop={int(best['population_size'])}, "
            f"gen={int(best['generations'])}, "
            f"mut=({best['early_mutation_rate']}/{best['mid_mutation_rate']}/{best['late_mutation_rate']}), "
            f"stagnant={int(best['max_stagnant'])}, "
            f"mean_dbi={best['mean_dbi_after_ga']}, "
            f"mean_silhouette={best['mean_silhouette']}"
        )


if __name__ == "__main__":
    main()
